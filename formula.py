import openpyxl


path = 'C:/PyTask/ExcelSheets/Apple-model-dummy (1).xlsx'

wb_formulas = openpyxl.load_workbook(path, data_only=False)
ws_formulas = wb_formulas.active


wb_values = openpyxl.load_workbook(path, data_only=True)
ws_values = wb_values.active


for row in ws_formulas.iter_rows(min_row=1, max_row=ws_formulas.max_row):
    for cell in row:

        if isinstance(cell, openpyxl.cell.MergedCell):
            continue
        
        if cell.column_letter in ['B']: 
            formula = cell.value if isinstance(cell.value, str) and cell.value.startswith('=') else "No formula"
            value = ws_values[cell.coordinate].value
            
            print(f"{formula}")
