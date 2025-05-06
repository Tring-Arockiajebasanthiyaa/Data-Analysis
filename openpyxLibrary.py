import re
import openpyxl
import json
import numpy as np

from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter


path = 'C:/PyTask/ExcelSheets/Apple-model-dummy (1).xlsx'
path1='C:\PyTask\ExcelSheets\Financial_Report (17).xlsx'
path2='C:\PyTask\ExcelSheets\Financial_Report Apple.xlsx'

wb=openpyxl.load_workbook(path , data_only=True)
wb1=openpyxl.load_workbook(path1,data_only=True)
wb2=openpyxl.load_workbook(path2,data_only=True)

name="CONDENSED CONSOLIDATED STATEMEN"

ws=wb.active
col=ws.max_column
rows=ws.max_row
print("\nColumns:",col ,"\nRow:",rows)
for i in range(6, rows+1):
    cell_obj=ws.cell(row=i,column=1).value
    cell_obj1=ws.cell(row=i,column=2).value
    print("\n",cell_obj,"\t\t\t\t\t\t\t",cell_obj1)
    

#Q2
ws2=wb1[name]

values=[]
print(ws2.title)

headings = []  
sub_key=[]
keys=""
values=[]
for row in ws2.iter_rows(min_row=3):  
    first_cell = row[0] 
    other_cells = row[1:]  
    
    if first_cell.value and all(
        (cell.value is None) or (str(cell.value).strip() == '') or (str(cell.value).strip() == '\u00a0')
        for cell in other_cells
    ):
        keys=first_cell.value
        headings.append(first_cell.value) 
        
    else:
        sub_key.append(first_cell.value + "_" + keys)
        values.append(row[1].value)

dict1=dict(zip(sub_key,values))
print(json.dumps(dict1,indent=4))

#Q3

ws3=wb2[name]
print(ws3.title)
heading1 = []  
sub_key1=[]
key1=""
value1=[]
for row in ws3.iter_rows(min_row=3):  
    first_cell = row[0] 
    other_cells = row[1:]  
    
    if first_cell.value and all(
        (cell.value is None) or (str(cell.value).strip() == '') or (str(cell.value).strip() == '\u00a0')
        for cell in other_cells
    ):
        key1=first_cell.value
        heading1.append(first_cell.value) 
        
    else:
        sub_key1.append(first_cell.value + "_" + key1)
        value1.append(row[1].value)

dict2=dict(zip(sub_key1,value1))
print(json.dumps(dict2,indent=4))

dict4={v:k for k , v in dict1.items()}



#create new Excel

# workbook=openpyxl.Workbook()
# new_sheet=workbook.create_sheet("Sheet2")
# workbook.save("Altered_excel.xlsx")

#create new column
# new_col = ws.max_column + 1

# for row in ws.iter_rows(min_row=9 , max_row=81):
#     search_row=row[1]
#     val=search_row.value
#     print(search_row.value)
    
#     if val not in ('', 'NA', None , 0 , 'REP' ,'-'):
#         q2_key = dict4.get(val)
#         q3_value = dict2.get(q2_key)
#         ws.cell(row=search_row.row, column=new_col).value = q3_value
        
# ws.cell(row=5 , column=3).value="Updated Values from Q3"     
# ws.cell(row=5, column=3).alignment = Alignment(wrap_text=True)
new_col = ws.max_column + 1
new_col_letter = get_column_letter(new_col)

# Add header
ws.cell(row=5, column=new_col).value = "Updated Values from Q3"
ws.cell(row=5, column=new_col).alignment = Alignment(wrap_text=True)


pattern = re.compile(r'(?<![A-Z])\$?B\$?(\d+)', re.IGNORECASE)


for row in ws.iter_rows(min_row=9, max_row=81):
    search_cell = row[1]  
    val = search_cell.value
    row_idx = search_cell.row
    target_cell = ws.cell(row=row_idx, column=new_col)

    
    if isinstance(val, str) and val.startswith('='):
        new_formula = pattern.sub(lambda m: f'{new_col_letter}{m.group(1)}', val)
        target_cell.value = new_formula

    
    elif val not in ('', 'NA', None, 0, 'REP', '-'):
        q2_key = dict4.get(val)
        q3_value = dict2.get(q2_key)
        target_cell.value = q3_value

wb.save(path)

for i in range(9,ws.max_row+1):
    col3_value=ws.cell(row=i,column=new_col).value
    print(col3_value)



